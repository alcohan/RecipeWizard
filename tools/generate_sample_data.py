"""Read RecipeBuilder XLSX and write sql/setup/sampledata.sql.

Run from anywhere: `python tools/generate_sample_data.py`
"""
import openpyxl
from datetime import datetime, date
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
PATH = r'C:\Users\Adrian\Downloads\RecipeBuilder 2025 JULY.xlsx'
OUT = PROJECT_ROOT / 'sql' / 'setup' / 'sampledata.sql'


def sql_str(s):
    if s is None or s == '':
        return 'NULL'
    s = str(s).rstrip()
    return "'" + s.replace("'", "''") + "'"


def sql_num(n, default='0'):
    if n is None:
        return default
    if isinstance(n, float) and n != n:
        return default
    if isinstance(n, float):
        if n.is_integer():
            return str(int(n))
        return repr(n)
    return str(n)


def sql_date(d):
    if d is None:
        return 'NULL'
    if isinstance(d, (datetime, date)):
        return "'" + d.strftime('%Y-%m-%d') + "'"
    return "'" + str(d) + "'"


wb = openpyxl.load_workbook(PATH, data_only=True)

sheet = wb['Ingredients & Recipes']
HEADER_ROW = 3
DATA_START = 4

headers = list(next(sheet.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True)))
recipes_label_col = headers.index('Recipes')
templates_label_col = headers.index('Templates')
recipe_names = [h for h in headers[recipes_label_col + 1:templates_label_col] if h]
print(f'{len(recipe_names)} recipes')

ingredients = []
for row_idx, row in enumerate(sheet.iter_rows(min_row=DATA_START, max_row=sheet.max_row, values_only=True), start=DATA_START):
    name = row[0]
    if name is None or str(name).strip() == '':
        continue
    ingredients.append({
        'id': len(ingredients) + 1,
        'name': str(name).strip(),
        'category': row[1],
        'unit': row[2],
        'portion': row[3],
        'weight': row[4],
        'cost': row[5],
        'calories': row[6],
        'ttlfat': row[7],
        'satfat': row[8],
        'cholesterol': row[9],
        'sodium': row[10],
        'carb': row[11],
        'fiber': row[12],
        'sugar': row[13],
        'protein': row[14],
        'allergens': {
            'Meat': row[15], 'Coconut': row[16], 'Fish': row[17], 'Shellfish': row[18],
            'Dairy': row[19], 'Eggs': row[20], 'Gluten': row[21], 'Tree Nuts': row[22],
            'Peanuts': row[23], 'Soy': row[24], 'Sesame': row[25],
        },
        'recipe_qtys': dict(zip(recipe_names, row[recipes_label_col + 1:templates_label_col])),
    })

print(f'{len(ingredients)} ingredients')
name_to_id = {ing['name']: ing['id'] for ing in ingredients}

connections = []
for ing in ingredients:
    for rname, qty in ing['recipe_qtys'].items():
        if qty is None:
            continue
        if isinstance(qty, float) and qty != qty:
            continue
        rid = recipe_names.index(rname) + 1
        connections.append((rid, None, ing['id'], qty))
print(f'{len(connections)} connections')

categories = sorted({ing['category'] for ing in ingredients if ing['category']})
tag_id_map = {tag: i + 1 for i, tag in enumerate(categories)}
print(f'{len(categories)} tags (categories only — allergens now go in their own table)')

ing_tags = []
for ing in ingredients:
    if ing['category']:
        ing_tags.append((tag_id_map[ing['category']], ing['id']))

allergen_names = ['Meat', 'Coconut', 'Fish', 'Shellfish', 'Dairy', 'Eggs', 'Gluten', 'Tree Nuts', 'Peanuts', 'Soy', 'Sesame']
allergen_id_map = {name: i + 1 for i, name in enumerate(allergen_names)}
ing_allergens = []
for ing in ingredients:
    for allergen, flag in ing['allergens'].items():
        if flag:
            ing_allergens.append((allergen_id_map[allergen], ing['id']))
print(f'{len(allergen_names)} allergens, {len(ing_allergens)} ingredient_allergens links')

ph_sheet = wb['Price History']
# Unit conversion: case_yield in yd_unit -> portion_unit. Derived from the workbook's Units sheet.
CONVERSIONS = {
    ('lb', 'g'): 454,
    ('gal', 'f.oz'): 128,
    ('gal', 'g'): 3632,
    ('w.oz', 'g'): 28.375,
    ('g', 'g'): 1,
    ('ea', 'ea'): 1,
    ('lb', 'lb'): 1,
    ('ea', '1/6 wedge'): 6,
}
# Undated rows in the workbook represent the original baseline data import —
# treat them as the oldest entry by stamping them with the earliest dated
# effective_date found in the sheet. Any later dated row wins via the
# update_ingredient_price trigger's `ORDER BY effective_date DESC` tiebreaker.
ph_rows = list(ph_sheet.iter_rows(min_row=2, values_only=True))
dated_values = [r[1] for r in ph_rows if r and len(r) > 1 and r[1] is not None]
ORIGINAL_DATE = min(dated_values).date() if dated_values else date.today()
prices = []
skipped_prices = 0
for row in ph_rows:
    if len(row) < 11:
        skipped_prices += 1
        continue
    ing_name, dt, _portion_cost, _pack, case_price, case_yield, yd_unit, yd_pct, portion, portion_unit, notes = row[:11]
    if not ing_name or ing_name not in name_to_id:
        skipped_prices += 1
        continue
    if dt is None:
        dt = ORIGINAL_DATE
    if case_price is None or case_yield is None or not portion:
        skipped_prices += 1
        continue
    factor = CONVERSIONS.get((yd_unit, portion_unit))
    if factor is None:
        skipped_prices += 1
        continue
    units_per_case = case_yield * factor * (yd_pct if yd_pct else 1) / portion
    prices.append({
        'ingredient_id': name_to_id[ing_name],
        'units_per_case': units_per_case,
        'case_price': case_price,
        'date': dt,
        'notes': notes,
    })
print(f'{len(prices)} price rows ({skipped_prices} skipped)')

lines = []
lines.append('-- Generated from "RecipeBuilder 2025 JULY.xlsx"')
lines.append(f'-- {len(ingredients)} ingredients, {len(recipe_names)} recipes, {len(connections)} connections, {len(prices)} price rows')
lines.append('')

lines.append('INSERT INTO Ingredients Values')
ing_rows = []
for ing in ingredients:
    ing_rows.append(
        f"({ing['id']},{sql_str(ing['name'])},{sql_str(ing['unit'])},{sql_str(ing['portion'])},"
        f"{sql_num(ing['weight'])},{sql_num(ing['cost'])},{sql_num(ing['calories'])},"
        f"{sql_num(ing['ttlfat'])},{sql_num(ing['satfat'])},{sql_num(ing['cholesterol'])},"
        f"{sql_num(ing['sodium'])},{sql_num(ing['carb'])},{sql_num(ing['fiber'])},"
        f"{sql_num(ing['sugar'])},{sql_num(ing['protein'])})"
    )
lines.append('\n,'.join(ing_rows))
lines.append(';')
lines.append('')

lines.append('INSERT INTO Recipes Values')
rec_rows = [f"({i},{sql_str(rname)},'each',1)" for i, rname in enumerate(recipe_names, start=1)]
lines.append('\n,'.join(rec_rows))
lines.append(';')
lines.append('')

lines.append('INSERT INTO CONNECTIONS Values')
conn_rows = [f"({p},NULL,{c},{sql_num(q)})" for (p, _, c, q) in connections]
lines.append('\n,'.join(conn_rows))
lines.append(';')
lines.append('')

lines.append("INSERT INTO suppliers (name) VALUES ('Sysco'),('Charlies');")
lines.append('')

if categories:
    lines.append('INSERT INTO tags (name) VALUES')
    lines.append(',\n'.join(f"({sql_str(t)})" for t in categories))
    lines.append(';')
    lines.append('')

if ing_tags:
    lines.append('INSERT INTO ingredient_tags_mapping (tag_id, ingredient_id) VALUES')
    lines.append(',\n'.join(f"({t},{i})" for t, i in ing_tags))
    lines.append(';')
    lines.append('')

lines.append('INSERT INTO allergens (name, sortOrder) VALUES')
lines.append(',\n'.join(f"({sql_str(name)},{i + 1})" for i, name in enumerate(allergen_names)))
lines.append(';')
lines.append('')

if ing_allergens:
    lines.append('INSERT INTO ingredient_allergens (allergen_id, ingredient_id) VALUES')
    lines.append(',\n'.join(f"({a},{i})" for a, i in ing_allergens))
    lines.append(';')
    lines.append('')

if prices:
    lines.append('INSERT INTO ingredient_prices')
    lines.append('    (ingredient_id, units_per_case, case_price, effective_date, notes)')
    lines.append('VALUES')
    price_rows = [
        f"({p['ingredient_id']},{p['units_per_case']},{p['case_price']},{sql_date(p['date'])},{sql_str(p['notes'])})"
        for p in prices
    ]
    lines.append('\n,'.join(price_rows))
    lines.append(';')

output = '\n'.join(lines)

with open(OUT, 'w', encoding='utf-8') as f:
    f.write(output)

print(f'\nWrote {len(output):,} chars to {OUT}')
